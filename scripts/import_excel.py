from openpyxl import load_workbook
import sqlite3


print("=" * 60)
print("OPENING DATABASE")
print("=" * 60)

connection = sqlite3.connect("database/prl_furniture.db")
cursor = connection.cursor()


print("=" * 60)
print("LOADING EXCEL")
print("=" * 60)

workbook = load_workbook(
    "data/furniture.xlsx",
    data_only=True,
)

print("Sheets:", workbook.sheetnames)

sheet = workbook["Katalog"]

print("Max row:", sheet.max_row)
print("Max column:", sheet.max_column)


cursor.execute("DELETE FROM furniture")


count = 0

for row in sheet.iter_rows(min_row=2, values_only=True):

    print("-" * 40)
    print(row)

    (
        _,
        model,
        common_name,
        reference_id,
        designer,
        production_years,
        category,
        manufacturer,
        visual_features,
        construction_features,
        similar_models,
        model_family,
        sources,
        confidence,
        search_features,
    ) = row

    cursor.execute(
        """
        INSERT INTO furniture (

            model,
            common_name,
            reference_id,

            designer,
            production_years,
            category,
            manufacturer,

            visual_features,
            construction_features,

            similar_models,
            model_family,

            sources,

            confidence,

            search_features

        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            model,
            common_name,
            reference_id,

            designer,
            production_years,
            category,
            manufacturer,

            visual_features,
            construction_features,

            similar_models,
            model_family,

            sources,

            confidence,

            search_features,
        ),
    )

    count += 1


connection.commit()

cursor.execute("SELECT COUNT(*) FROM furniture")
database_count = cursor.fetchone()[0]

connection.close()

print("=" * 60)
print("IMPORT FINISHED")
print("=" * 60)
print("Imported rows:", count)
print("Rows in database:", database_count)
print("=" * 60)